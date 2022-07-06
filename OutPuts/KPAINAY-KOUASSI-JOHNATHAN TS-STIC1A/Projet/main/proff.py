import sys
import PyQt5
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import sqlite3
from PyQt5.uic import loadUi
from PyQt5.uic import loadUiType
import xlrd, openpyxl
from openpyxl import workbook, load_workbook

prof = loadUiType("C:\\Users\\lawre\\OneDrive - INP-HB\\Desktop\LOKI\\Projet\\professeur.ui")

database = sqlite3.connect("gnote_base.db")
cursor = database.cursor()

class professeur(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        loadUi("C:\\Users\\lawre\\OneDrive - INP-HB\\Desktop\LOKI\\Projet\\professeur.ui", self)

        self.filehandle = ""
        self.wb = ""
        ##########################################NAVIGATION####################################################
        self.pushButton.clicked.connect(self.saisir)
        self.pushButton_2.clicked.connect(self.export)
        self.pushButton_3.clicked.connect(self.mettreAJour)

        ################################OPERATIONS FICHIERS######################################################
        self.pushButton_16.clicked.connect(self.findFile)
        self.pushButton_9.clicked.connect(self.saveImport)
        self.pushButton_10.clicked.connect(self.annulerImport)
        self.pushButton_19.clicked.connect(self.emplace)
        self.pushButton_17.clicked.connect(self.annulerExport)
        self.pushButton_18.clicked.connect(self.saveExportPdf)
        self.pushButton_11.clicked.connect(self.saveExportExcel)
        self.pushButton_8.clicked.connect(self.saveNotes)
        self.pushButton_5.clicked.connect(self.clrNote)

        lis = QTableWidget()



        self.saveState()

    def saveNotes(self):
        try:

            valeurMax = 20
            Date = self.dateEdit.text()
            Coef = self.lineEdit.text()
            matiere = self.comboBox.currentText()
            matrEt = self.lineEdit_3.text()
            valeur = self.lineEdit_2.text()
            matrProf = self.lineEdit_4.text()
            numero = self.lineEdit_5.text()

            query = """INSERT INTO NOTE VALUES(?,?,?,?,?,?,?,?) """
        except:
            self.comfirmation3.setText(" Erreur, Assurez-vous que les informations fournies sont correctes")
        try:
            cursor.execute(query, (numero, valeur, valeurMax, Date, Coef, matrProf, matiere, matrEt,))
            self.comfirmation3.setText("SUCCESSFULLY WRITTEN")
            database.commit()
        except:
            self.comfirmation3.setText("Erreur!!")

        try:
            lib = "Vos notes ont étéés mises à jour"
            cursor.execute("""INSERT INTO Notification (id_notif, libéllé_notif, dete_notif, Matricule_etudiant) 
            VALUES(?,?,?,?)""", (numero, lib, Date, matrEt))
            database.commit()
        except:
            self.comfirmation3.setText("Erreur lors de l'envoie de la notification")

    def clrNote(self):
        pass
    def emplace(self):
        try:
            fname =QFileDialog.getExistingDirectory(self, 'Emplacement:', 'c\\')
            self.label_5.setText(str(fname))
        except:
            self.label_6.setText("Erreur!!")
    def listAction(self, clickedItem):
        try:
            for i in cursor.execute("SELECT * FROM ").fetchall():
                pass
        except:
            self.label_6.setText("Erreur!!")
    def saveExportPdf(self):
        #TODO: CREATE AN EXCEL FILE THAT WILL IN TURN BE CONVERTED TO PDF
        pass

    def saveExportExcel(self):
        #TODO: CREATE AN EXCEL FILE AND SAVE IN THE PROVIDED LOCATION
        pass
    def annulerExport(self):
        self.label_5.clear()

    def findFile(self):
        txt = QTextBrowser()
        # Essayer de trouver le fichier xlsx, xls, xlw ou xla
        try:
            fname = QFileDialog.getOpenFileName(self, 'Choisir un fichier', 'c:\\',
                                                "Excel Files (*.xlsx *.xls, *.xlw, *.xla )")
            self.label_3.setText(str(fname[0]))
            self.filehandle = fname[0]
            print(self.filehandle)
            # si le fichier est importe avec succès,
            # essayer de l'ouvrir
            try:
                wb = load_workbook(self.filehandle)
                sheet = wb.active
                print(sheet)
                # si il a ete ouvert avec succès,
                # essayer d'ecrire le contenu sur un tableau sur l'ecran
                try:
                    tab = QTableWidget()
                    table = QTableView()

                    self.tableWidget_2.setRowCount(200)
                    self.tableWidget_2.setColumnCount(200)
                    # boucle pour ecrire le contenu du fichier xlsx
                    for i in range(1, sheet.max_row + 1):
                        for j in range(1, sheet.max_column + 1):
                            # item =QTableWidgetItem("Note")
                            item = sheet.cell(i, j)
                            cel = QLabel(str(item.value))
                            # verifier si le champs est vide ou non
                            if cel.text() == "":
                                self.tableWidget_2.setCellWidget(i, j, "")
                            else:
                                self.tableWidget_2.setCellWidget(i, j, cel)


                except:
                    # si il y a eu une erreur lors de l'ecriture des contenus,
                    # afficher un message d'erreur
                    print("Error writing to textView!")
                    self.label_4.setText("Une Erreur s'est prosuite")
            except:
                # si il y a eu erreur lors de l'ouverture du ficher,
                # afficher un message d'erreur
                print("Excel Module Error!")
                self.label_4.setText("Une Erreur s'est produite")

        except:
            print("Error finding file!")
            self.label_4.setText("Erreur lors de l'ouverture du Fichier ")

    def saveImport(self):
        # read data in the table
        try:
            self.wb = load_workbook(self.filehandle)
            self.wb.save("workbook.xlsx")
        except:
            print("Error!")

        # execute a query for each cell or execute queries by rows?

        pass


    def annulerImport(self):
        self.label_3.setText("")
        self.tableWidget_2.clear()

    def export(self):
        self.stackedWidget.setCurrentWidget(self.page_2)

    def saisir(self):
        self.stackedWidget.setCurrentWidget(self.page)
        try:
            for j in cursor.execute(""" SELECT Nom_de_la_matiere FROM MATIERE""").fetchall():
                self.comboBox.addItems(j)
        except:
            self.comfirmation3.setText("Erreur!!")

        self.tableWidget_5.clearContents()
        try:
            for l in cursor.execute("SELECT nom_de_la_classe_ FROM CLASSE").fetchall():
                self.comboBox_3.addItems(l)
        except:
            self.comfirmation3.setText("ERREUR à l'initialisation")
        clas = self.comboBox_3.currentText()
        try:
            data = cursor.execute("""SELECT Nom_etudiant_, Prenom_etudiant_ FROM ETUDIANT WHERE nom_de_la_classe_=?""",
                                  (clas,)).fetchall()
        except:
            print("Error getting data")

        try:
            self.tableWidget_5.setRowCount(len(data))
            self.tableWidget_5.setColumnCount(len(data[0]))
            print(data)
            for i in range(len(data)):
                for j in range(len(data[0])):
                    cell = QLabel(str(data[i][j]))
                    self.tableWidget_5.setCellWidget(i, j, cell)
        except:
            print("Error Writing Data")

    def listStudents(self):
        pass

    def mettreAJour(self):
        self.stackedWidget.setCurrentWidget(self.page_3)
        #TODO: SET UP THE WORKSPACE FOR MODIFICATIONS

def main():
    app = QApplication(sys.argv)

    prof = professeur()
    prof.show()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
